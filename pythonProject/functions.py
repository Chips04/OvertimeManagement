import datetime
from openpyxl import load_workbook
import re


# fillTime 填充【值班日期】和【值班日期日期格式】
# n: 原有表格总共几行
# m: 期望填充到表格到第几行
def ft(df, n, m):
    n = n - 1
    for i in range(m - n - 1):
        if df.loc[i + n, '值班组名'] == df.loc[i - 1 + n, '值班组名']:
            df.loc[i + n, '值班日期日期格式'] = df.loc[i - 1 + n, '值班日期日期格式']
        else:
            df.loc[i + n, '值班日期日期格式'] = df.loc[i - 1 + n, '值班日期日期格式'] + datetime.timedelta(days=1)
        df.loc[i + n, '值班日期'] = df.loc[i + n, '值班日期日期格式'].strftime('%Y-%m-%d')
        df.loc[i + n, 'currentMonthFlag'] = 0


# replaceName 重名人员名字替换为人员id
def rn(df, cname):
    for index, item in enumerate(df[cname]):
        if type(item) == str:
            item = item.replace('张华', '#szzl_zhanghua#')
            item = item.replace('陈燕珊', '#szzl_chenyanshan2#')
            item = item.replace('陈燕', '#szzl_chenyan18#')
            item = item.replace('吴晓燕', '#szzl_wuxiaoyan2#')
            item = item.replace('张远华', '#szzl_zhangyuanhua2#')
            item = item.replace('王莹', '#szzl_wangying#')
            item = item.replace('常春艳', '#mx93#')
            item = item.replace('林泽鑫', '#bzsq_13#')
            df.loc[index, cname] = item


# ReplaceName 重名人员名字替换为人员id openpyxl版
# file_name 要处理的文件路径
# 要替换的列列表 如['A', 'B', 'C']
def rn2(file_name, handle_name, columns_to_replace):
    wb = load_workbook(file_name)
    ws = wb.active
    # 替换规则
    replacement_rules = {
        '张华': '#szzl_zhanghua#',
        '陈燕珊': '#szzl_chenyanshan2#',
        '陈燕': '#szzl_chenyan18#',
        '吴晓燕': '#szzl_wuxiaoyan2#',
        '张远华': '#szzl_zhangyuanhua2#',
        '王莹': '#szzl_wangying#',
        '常春艳': '#mx93#',
        '林泽鑫': '#bzsq_13#'
    }
    # 正则表达式，匹配所有需要替换的姓名
    pattern = re.compile('|'.join(map(re.escape, replacement_rules.keys())))
    for column in columns_to_replace:
        for cell in ws[column]:
            if isinstance(cell.value, str) and pattern.search(cell.value):
                cell.value = pattern.sub(lambda mo: replacement_rules[mo.string[mo.start():mo.end()]], cell.value)
    wb.save(handle_name)


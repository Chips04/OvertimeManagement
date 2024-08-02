import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import Categorical
import os
import fnmatch
from openpyxl.styles import PatternFill
import json
import subprocess
import argparse

# 读取txt文件内容
# 读取文件内容，并手动处理可能的 UTF-8 BOM
with open('variables.txt', 'rb') as file:  # 使用二进制模式打开文件
    content = file.read()
    # 检查并去除 UTF-8 BOM（如果有的话）
    if content.startswith(b'\xef\xbb\xbf'):
        content = content[3:]  # 跳过 BOM
    content = content.decode('utf-8').strip()  # 解码为 UTF-8 并去除空白字符

# 尝试将内容转换为字典，注意这里假设了txt文件的内容是合法的JSON格式
try:
    data = json.loads(content)  # 将字符串转换为字典
except json.JSONDecodeError as e:
    print(f"无法解析txt文件内容为JSON: {e}")
    exit(1)

# 从字典中提取变量
path = data['path']

fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

# 打开现有的工作簿
wb_origin = openpyxl.load_workbook(path + '模板/2407附件2：加班费审批表 .xlsx')  # ！！！
ws_origin = wb_origin.worksheets[0]
# 工资基数
ws_salary = wb_origin.worksheets[1]
# 定义列名
column_names = ['学历', '工资基数']
# 将工作表的值转换为DataFrame，并指定列名
df_salary = pd.DataFrame(ws_salary.values, columns=column_names)
df_salary = df_salary.drop(0)
wb_origin2 = openpyxl.load_workbook(path + '模板/2407附件3：加班日志汇总表.xlsx')  # ！！！
ws_origin2 = wb_origin2.active
wb_origin3 = openpyxl.load_workbook(path + '模板/单日加班超4小时申请汇总表模板.xlsx')  # ！！！
ws_origin3 = wb_origin3.active
wb_origin4 = openpyxl.load_workbook(path + '模板/2407附件1：加班费申报表.xlsx')  # ！！！
ws_origin4 = wb_origin4.active
wb_origin5 = openpyxl.load_workbook(path + '模板/2407附件2：加班费审批表 .xlsx')  # ！！！
ws_origin5 = wb_origin5.active
wb_origin6 = openpyxl.load_workbook(path + '模板/2407附件2：加班费审批表 .xlsx')  # ！！！
ws_origin6 = wb_origin6.active
wb_origin7 = openpyxl.load_workbook(path + '模板/240729附件1：加班费申报表（第二季度）.xlsx')  # ！！！
ws_origin7 = wb_origin7.active

# 获取数据源dataframe
df1 = pd.DataFrame()
# 指定目录和通配符
directory = path
pattern = '加班申请*.xlsx'
# 获取匹配的文件列表
matches = []
for filename in os.listdir(directory):
    if fnmatch.fnmatch(filename, pattern):
        matches.append(os.path.join(directory, filename))

# 打开第一个匹配的文件并读取内容
if matches:
    df1 = pd.read_excel(matches[0])
else:
    print('No matching files found.')


# # 获取数据源dataframe
# df2 = pd.DataFrame()

pattern2 = '社区工作人员花名册*.xlsx'
# 获取匹配的文件列表
matches2 = []
for filename2 in os.listdir(directory):
    if fnmatch.fnmatch(filename2, pattern2):
        matches2.append(os.path.join(directory, filename2))

# # 打开第一个匹配的文件并读取内容
# if matches2:
#     df2 = pd.read_excel(matches2[0])
# else:
#     print('No matching files found2.')

# 获取数据源dataframe
df14 = pd.DataFrame()

pattern14 = '花名册_各种版本*.xlsx'
# 获取匹配的文件列表
matches14 = []
for filename14 in os.listdir(directory):
    if fnmatch.fnmatch(filename14, pattern14):
        matches14.append(os.path.join(directory, filename14))

# 打开第一个匹配的文件并读取内容
if matches14:
    df14 = pd.read_excel(matches14[0])
else:
    print('No matching files found14.')

# print(df14)

# 获取数据源dataframe
df3 = pd.DataFrame()
# 指定目录和通配符
pattern3 = '人员管理-请休假*.xlsx'
# 获取匹配的文件列表
matches3 = []
for filename3 in os.listdir(directory):
    if fnmatch.fnmatch(filename3, pattern3):
        matches3.append(os.path.join(directory, filename3))

# 打开第一个匹配的文件并读取内容
if matches3:
    df3 = pd.read_excel(matches3[0])
else:
    print('No matching files found.')


# df2 = df2.rename(columns={'工作人员姓名': '姓名', '人员类别（加班费使用）': '人员类别', '学历（加班费使用）': '学历'})
df14 = df14.rename(columns={'工作人员姓名': '姓名', '人员类别（加班费使用）': '人员类别', '学历（加班费使用）': '学历'})


# 获得年月
max_value = df1['加班日期'].max()
current_year = max_value.year
current_month = max_value.month
month1 = 0
# 根据月份确定季度
if 1 <= current_month <= 3:
    current_quarter = '一'
    month1 = 1
elif 4 <= current_month <= 6:
    current_quarter = '二'
    month1 = 4
elif 7 <= current_month <= 9:
    current_quarter = '三'
    month1 = 7
else:
    current_quarter = '四'
    month1 = 10
month2 = month1 + 1
month3 = month1 + 2

# 只保留报加班费的行
df_log = df1.copy()
# df1 = df1[df1['报加班费还是给调休'] == '报加班费']
df1 = df1[df1['报加班费还是给调休'].str.contains('加班费')]

# 删除不同意的行
if '初核意见' in df1.columns:
    df1 = df1[df1['初核意见'] != '不同意']
    df1 = df1[df1['复核意见'] != '不同意']
    df1 = df1[df1['初初核意见'] != '不同意']

# 合并df1和df2
# 确保加班日期是datetime类型
df1['加班日期'] = pd.to_datetime(df1['加班日期'])

# 为 df1 和 df14 分别添加新的列，只包含年月的部分
df1['加班日期_年月'] = df1['加班日期'].dt.to_period('M')
df14['名单生效年月_年月'] = df14['名单生效年月'].dt.to_period('M')

# 使用新添加的列进行合并
merged_df1_2 = pd.merge(df1, df14,
                        left_on=['姓名', '加班日期_年月'],
                        right_on=['姓名', '名单生效年月_年月'],
                        how='left')

# 如果您不再需要额外的年月列，可以在合并后删除它们
# merged_df1_2.drop(['加班日期_年月', '名单生效年月_年月'], axis=1, inplace=True)

# 合并merged_df1_2和df3
result = pd.merge(merged_df1_2, df_salary, on='学历', how='left')
# 排序
result = result.sort_values(by=['编号', '加班开始时间'], ascending=[True, True])

# 定义一个函数，该函数接收DataFrame的行作为输入，并返回你想要添加到新列的值
def calculate_value(row):
    # 在这里，你可以根据'A'和'B'列的值来计算新列的值
    # return row['A'] + row['B']
    G4 = 0
    if row['加班类型'] == '工作日加班':
        G4 = 1.5
    elif row['加班类型'] == '周末加班':
        G4 = 2
    elif row['加班类型'] == '法定节假日加班':
        G4 = 3
    J4 = row['工资基数']/21.75/8
    return round(G4 * J4 * row['加班计算小时数'], 2)

# 使用apply函数，指定axis=1以在行上应用函数
result['加班费金额'] = result.apply(calculate_value, axis=1)

result = result.reset_index(drop=True)
month1_df = result[(result['名单生效年月_年月'].dt.year == current_year) & (result['名单生效年月_年月'].dt.month == month1)]
month2_df = result[(result['名单生效年月_年月'].dt.year == current_year) & (result['名单生效年月_年月'].dt.month == month2)]
month3_df = result[(result['名单生效年月_年月'].dt.year == current_year) & (result['名单生效年月_年月'].dt.month == month3)]

# 结果的表格
columns = ['序号', '人员类别', '姓名', '="加班日期"&CHAR(10)&"附时间段"', '加班事项类型', '加班事由', '加班类型', '加班费倍率', '学历', '工资基数', '加班费基数/小时', '加班时长（小时）', '加班费金额（元）', '个人小计']
columns2 = ['序号', '人员类别', '姓名', '="加班日期"&CHAR(10)&"附时间段"', '加班事项类型', '加班事由', '加班类型', '加班费倍率', '学历', '工资基数', '加班费基数/小时', '加班时长（小时）', '加班费金额（元）', '个人小计', '备注', '画×', '报加班费还是给调休', '本次加班获得调休时长', 'data_id_x']
df_final = pd.DataFrame(columns=columns)


# 数据处理
# begin_num_final: 数据开始行数, 加班申请表：4，超过4小时表： 3
# letter_final: 公式列开始列数, 加班申请表：F，超过4小时表： E
# flag: 加班审批表分公司：1 加班审批表：0 新加班审批表：3
def handel_df_final(begin_num_final, letter_final, flag):
    current_company = ''
    current_num0 = 0
    result_handeled0 = result
    result_handeled = result
    result_handeled['类别（加班费使用）'] = Categorical(result_handeled['类别（加班费使用）'], categories=['直聘', '劳务派遣（天域）', '劳务派遣（天悦）', '劳务派遣（翰林晟）', '未过渡', '党务', '其他'], ordered=True)
    result_handeled = result_handeled.sort_values(by=['类别（加班费使用）', '编号'])
    result_handeled = result_handeled.reset_index(drop=True)
    # 超过4小时表
    if begin_num_final == 3:
        # 删除工作日加班的行
        result_handeled1 = result_handeled0[result_handeled0['加班类型'] != '工作日加班']
        # 删除加班时间小于等于4的行
        result_handeled = result_handeled1[result_handeled1['加班计算小时数'] > 4]
    # 新加班审批表从第5行开始
    begin_num_final2 = 5 if flag == 3 else begin_num_final
    df_final_handeled = pd.DataFrame(columns=columns)
    row_index = 0
    row_name = ''
    result_handeled = result_handeled.reset_index()

    # 检查指定索引的行是否存在
    if 0 < result_handeled.shape[0]:
        current_company = result_handeled.iloc[0]['类别（加班费使用）']
    else:
        print("DataFrame 是空的，无法访问索引为0的行")
    for index_final, row_final in result_handeled.iterrows():
        # 将字符串转换为datetime对象
        start_datetime = row_final['加班开始时间']
        end_datetime = row_final['加班结束时间']
        # 格式化输出
        start_formatted = start_datetime.strftime("%Y.%m.%d")
        start_formatted2 = start_datetime.strftime("%H:%M")
        end_formatted = end_datetime.strftime("%H:%M")

        if not pd.isna(row_final['加班开始时间2']):
            start_datetime2 = row_final['加班开始时间2']
            start_formatted4 = start_datetime2.strftime("%H:%M")
        if not pd.isna(row_final['加班结束时间2']):
            end_datetime2 = row_final['加班结束时间2']
            end_formatted2 = end_datetime2.strftime("%H:%M")
        if not pd.isna(row_final['加班开始时间3']):
            start_datetime3 = row_final['加班开始时间3']
            start_formatted6 = start_datetime3.strftime("%H:%M")
        if not pd.isna(row_final['加班结束时间3']):
            end_datetime3 = row_final['加班结束时间3']
            end_formatted3 = end_datetime3.strftime("%H:%M")

        if row_final['你这一天分了几段时间'] == '一段':
            # 构造最终格式的字符串
            time1 = f'{start_formatted}\n（{start_formatted2}-{end_formatted}）'
        elif row_final['你这一天分了几段时间'] == '两段':
            # 构造最终格式的字符串
            time1 = f'{start_formatted}\n（{start_formatted2}-{end_formatted}）\n（{start_formatted4}-{end_formatted2}）'
        else:
            # 构造最终格式的字符串
            time1 = f'{start_formatted}\n（{start_formatted2}-{end_formatted}）\n（{start_formatted4}-{end_formatted2}）\n（{start_formatted6}-{end_formatted3}）'
        if row_name != row_final['姓名']:
            row_name = row_final['姓名']
            row_index = row_index + 1
        if flag == 3:
            if row_final['类别（加班费使用）'] != current_company:
                current_num0 += 1
                current_company = row_final['类别（加班费使用）']
                row_index = 1
            newRow1 = pd.Series(
                [
                    row_index,
                    row_final['人员类别'],
                    row_final['姓名'],
                    time1,
                    row_final['加班事项类型'],
                    row_final['具体事由'],
                    "公休日加班" if row_final['加班类型'] == '周末加班' else row_final['加班类型'],
                    f'=IF({letter_final}{index_final + begin_num_final2 + current_num0}="","",VLOOKUP({letter_final}{index_final + begin_num_final2 + current_num0},{{"工作日加班",1.5;"公休日加班",2;"法定节假日加班",3}},2,))',
                    row_final['学历'],
                    f'=VLOOKUP(I{index_final + begin_num_final2 +current_num0},工资基数!A:B,2,FALSE)',
                    f'=J{index_final + begin_num_final2 + current_num0}/21.75/8',
                    row_final["加班计算小时数"],
                    f'=ROUND(H{index_final + begin_num_final2 +current_num0}*K{index_final + begin_num_final2 + current_num0}*L{index_final + begin_num_final2 + current_num0},2)',
                    row_final['类别（加班费使用）'],
                    '',
                    '',
                    f'=IF(O{index_final + begin_num_final2 + current_num0}="","报加班费","给调休")',
                    f'=IF(O{index_final + begin_num_final2 + current_num0}="",0,K{index_final + begin_num_final2 + current_num0})',
                    row_final['data_id_x'],
                ],
                index=columns2
            )
        else:
            newRow1 = pd.Series(
                [
                    row_index,
                    row_final['人员类别'],
                    row_final['姓名'],
                    time1,
                    row_final['具体事由'],
                    "公休日加班" if row_final['加班类型'] == '周末加班' else row_final['加班类型'],
                    f'=IF({letter_final}{index_final + begin_num_final2}="","",VLOOKUP({letter_final}{index_final + begin_num_final2},{{"工作日加班",1.5;"公休日加班",2;"法定节假日加班",3}},2,))',
                    row_final['学历'],
                    "=VLOOKUP(H:H,工资基数!$A$1:$B$59,2,FALSE)",
                    f'=I{index_final + 4}/21.75/8',
                    row_final["加班计算小时数"],
                    f'=ROUND(G{index_final + 4}*J{index_final + 4}*K{index_final + 4},2)',
                    "",
                ],
                index=columns
            )
        # df_final_handeled = df_final_handeled._append(newRow1, ignore_index=True)
        # df_final_handeled = pd.concat([df_final_handeled, pd.DataFrame([newRow1])], ignore_index=True)
        df_final_handeled = pd.concat([df_final_handeled, pd.DataFrame([newRow1])], ignore_index=True, sort=False)
    return df_final_handeled



# # 只拷贝格式
# def copy_format(source_cell, target_sheet, target_row, target_col):
#     target_cell = target_sheet.cell(row=target_row, column=target_col)
#     # target_cell.value = target_value
#     # 设置单元格格式
#     target_cell.fill = copy.copy(source_cell.fill)
#     if source_cell.has_style:
#         # sheet.column_dimensions[column_letter].width
#         target_cell._style = copy.copy(source_cell._style)
#         target_cell.font = copy.copy(source_cell.font)
#         target_cell.border = copy.copy(source_cell.border)
#         target_cell.fill = copy.copy(source_cell.fill)
#         target_cell.number_format = copy.copy(source_cell.number_format)
#         target_cell.protection = copy.copy(source_cell.protection)
#         target_cell.alignment = copy.copy(source_cell.alignment)
#
#
# # 只拷贝格式
# def copy_format2(source_cell, target_cell):
#     # 设置单元格格式
#     target_cell.fill = copy.copy(source_cell.fill)
#     if source_cell.has_style:
#         # sheet.column_dimensions[column_letter].width
#         target_cell._style = copy.copy(source_cell._style)
#         target_cell.font = copy.copy(source_cell.font)
#         target_cell.border = copy.copy(source_cell.border)
#         target_cell.fill = copy.copy(source_cell.fill)
#         target_cell.number_format = copy.copy(source_cell.number_format)
#         target_cell.protection = copy.copy(source_cell.protection)
#         target_cell.alignment = copy.copy(source_cell.alignment)


# 创建一个新的工作簿
wb = Workbook()
sheet = wb.active
# wb1 = Workbook()
# sheet1 = wb1.active
wb_temp = Workbook()
sheet_temp = wb_temp.active

# dff = dff.drop(0)
# 将数据写入工作表000
# for r0 in dataframe_to_rows(dff, index=False, header=True):
#     ws_origin.append(r0)


# 模板表格变长
# ws_origin.insert_rows(idx=5, amount=len(dff) - 1)


# ******************************1.加班审批表************************************
# dff = handel_df_final(4, 'F', 0)
# # 填充数据 加班审批表
# BBB = 4
# for iii in range(len(dff)):
#     for jjj in range(len(dff.columns)):
#         cell_value = dff.iat[iii, jjj]
#         if cell_value != '':
#             ws_origin.cell(iii + BBB, jjj + 1, value=cell_value)
#
# # 合并单元格，小计
# current_num = 4
# count = 0
# for k in range(len(dff)):
#     if k == 0:
#         continue
#     elif dff.iat[k, 0] != dff.iat[k - 1, 0]:
#         ws_origin.merge_cells(f'A{current_num}:A{current_num + count}')
#         ws_origin.merge_cells(f'B{current_num}:B{current_num + count}')
#         ws_origin.merge_cells(f'C{current_num}:C{current_num + count}')
#         ws_origin.merge_cells(f'M{current_num}:M{current_num + count}')
#         ws_origin[f'M{current_num}'] = f'=SUM(L{current_num}:L{current_num + count})'
#         if k == len(dff) - 1:
#             ws_origin[f'M{k + 4}'] = ws_origin[f'L{k + 4}'].value
#         current_num = current_num + count + 1
#         count = 0
#     elif k == len(dff) - 1:
#         ws_origin.merge_cells(f'A{current_num}:A{current_num + count + 1}')
#         ws_origin.merge_cells(f'B{current_num}:B{current_num + count + 1}')
#         ws_origin.merge_cells(f'C{current_num}:C{current_num + count + 1}')
#         ws_origin.merge_cells(f'M{current_num}:M{current_num + count + 1}')
#         ws_origin[f'M{current_num}'] = f'=SUM(L{current_num}:L{current_num + count + 1})'
#     else:
#         count += 1

# *********************2.加班日志汇总表********************
# 只保留加班费的行
df20 = df_log[df_log['报加班费还是给调休'].str.contains('加班费')].copy()
# 删除值班的行
# df20 = df_log
# df20 = df_log[df_log['报加班费还是给调休'] != '按上级规定需要安排值守的，以及办事处安排的应急值守、安全生产巡查值班、社会治安巡逻督察的，另行制定定额标准执行。']
df20['加班事项类型'] = df20['加班事项类型'].fillna('未知')
if len(df20) > 0:
    df22 = df20.groupby(['加班日期', '加班事项类型', '加班类型', '加班类别', '具体事由']).size().reset_index(name='人数')
    df23 = df20.groupby(['加班日期', '加班事项类型', '加班类型', '加班类别', '具体事由'])['姓名'].agg('、'.join).reset_index()
    sorted_result1 = df22.sort_values(by='加班日期')
    sorted_result2 = df23.sort_values(by='加班日期')
    # merged_df = pd.merge(df22, df23[['姓名', '加班日期', '具体事由']], on=['加班日期', '具体事由'])
    # 现在，我们可以根据相同的键合并这两个DataFrame
    merged_df = pd.merge(df22, df23,
                         on=['加班日期','加班事项类型', '加班类型', '加班类别', '具体事由'])
    merged_df['日期和总人数'] = merged_df.apply(lambda row: f'{row["加班日期"].month}月{row["加班日期"].day}日', axis=1)
    merged_df = merged_df.rename(columns={'具体事由': '加班事项'})
    merged_df = merged_df.drop('加班日期', axis=1)
    merged_df = merged_df.reindex(columns=['日期和总人数', '加班事项类型', '加班类型', '加班类别', '加班事项', '人数', '姓名'])
    merged_df = merged_df.drop(['加班类型', '加班类别'], axis=1)
    # 填充数据 加班日志汇总表
    b2 = 5
    for i2 in range(len(merged_df)):
        for j2 in range(len(merged_df.columns)):
            cell_value2 = merged_df.iat[i2, j2]
            ws_origin2.cell(i2 + b2, j2 + 1, value=cell_value2)

    # 合并单元格 日期和总人数
    current_num2 = 5
    count2 = 0
    people_num = 0
    for k2 in range(len(merged_df)):
        if k2 == 0:
            people_num = ws_origin2[f"D{current_num2}"].value
        elif merged_df.iat[k2, 0] != merged_df.iat[k2 - 1, 0]:
            ws_origin2.merge_cells(f'A{current_num2}:A{current_num2 + count2}')
            ws_origin2[f'A{current_num2}'] = f'="{ws_origin2[f"A{current_num2}"].value}\n加班{people_num}人"'
            if k2 == len(merged_df) - 1:
                ws_origin2[
                    f'A{current_num2 + count2 + 1}'] = f'="{ws_origin2[f"A{current_num2 + count2 + 1}"].value}\n加班{merged_df.iat[k2, 3]}人"'
            current_num2 = current_num2 + count2 + 1
            count2 = 0
            people_num = merged_df.iat[k2, 3]

        elif k2 == len(merged_df) - 1:
            ws_origin2.merge_cells(f'A{current_num2}:A{current_num2 + count2 + 1}')
            ws_origin2[
                f'A{current_num2}'] = f'="{ws_origin2[f"A{current_num2}"].value}\n加班{people_num + merged_df.iat[k2, 3]}人"'
        else:
            count2 += 1
            people_num += merged_df.iat[k2, 3]
# ws_origin2['A200'] = "加班总人数：" + str(len(df20['姓名'].unique()))
# ws_origin2['H200'] = "申报加班费总时长：" + str(df20[df20['报加班费还是给调休'].str.contains('加班费')]['加班计算小时数'].sum())

# ******************************3.单日加班超4小时申请汇总表************************************
# df300 = handel_df_final(3, 'E', 0)
#
# # 设置列
# df31 = df300[['序号', '姓名', '="加班日期"&CHAR(10)&"附时间段"', '加班事由', '加班类型', '加班费倍率', '加班时长（小时）']]
# # 填充数据 单日加班超4小时申请汇总表
# cc = 3
# for i3 in range(len(df31)):
#     for j3 in range(len(df31.columns)):
#         cell_value = df31.iat[i3, j3]
#         if cell_value != '':
#             ws_origin3.cell(i3 + cc, j3 + 1, value=cell_value)
#
# # 合并单元格
# current_num3 = 3
# count3 = 0
# for k in range(len(df31)):
#     if k == 0:
#         continue
#     elif df31.iat[k, 0] != df31.iat[k - 1, 0]:
#         ws_origin3.merge_cells(f'A{current_num3}:A{current_num3 + count3}')
#         ws_origin3.merge_cells(f'B{current_num3}:B{current_num3 + count3}')
#         current_num3 = current_num3 + count3 + 1
#         count3 = 0
#     elif k == len(df31) - 1:
#         ws_origin3.merge_cells(f'A{current_num3}:A{current_num3 + count3 + 1}')
#         ws_origin3.merge_cells(f'B{current_num3}:B{current_num3 + count3 + 1}')
#     else:
#         count3 += 1

# *********************4.加班费申报表********************
df40 = result
# 定义一个函数，该函数接收DataFrame的行作为输入，并返回你想要添加到新列的值
def calculate_value(row):
    G4 = 0
    if row['加班类型'] == '工作日加班':
        G4 = 1.5
    elif row['加班类型'] == '周末加班':
        G4 = 2
    elif row['加班类型'] == '法定节假日加班':
        G4 = 3
    J4 = row['工资基数']/21.75/8
    return round(G4 * J4 * row['加班计算小时数'], 2)


# 使用apply函数，指定axis=1以在行上应用函数
df40['加班费金额'] = df40.apply(calculate_value, axis=1)
df41 = df40.groupby(['姓名', '人员类别', '加班费申报表顺序', '审批人（加班费使用）', '工资基数'])[['加班计算小时数', '加班费金额']].sum().reset_index()
# df41 = df40.groupby(['加班类型', '姓名', '人员类别', '加班费申报表顺序', '审批人（加班费使用）', '工资基数'])['加班计算小时数', '加班费金额'].sum().reset_index()
df42 = df41.sort_values('加班费申报表顺序', ascending=True).reset_index()
df42 = df42.drop(['加班费申报表顺序', "加班计算小时数", '审批人（加班费使用）'], axis=1)


df42 = df42.reindex(columns=['index', "人员类别", "姓名", '加班费金额'])
df42['index'] = range(1, len(df42) + 1)
# num44 = -1
# last_name = ""
# num_list = []
# for r_temp in dataframe_to_rows(df42, index=False, header=True):
#     if r_temp[1] != last_name:
#         last_name = r_temp[1]
#         num44 += 1
#     r_temp.insert(0, num44)
#     num_list.append(r_temp[0])
#     sheet_temp.append(r_temp)
# new_column = pd.Series(num_list, name='序号')
# new_column.pop(0)
# df42 = df42.reset_index(drop=True)
# new_column = new_column.reset_index(drop=True)
# df42 = pd.concat([new_column, df42], axis=1)
# df42['加班类型'] = df42['加班类型'].replace('周末加班', '公休日加班')
# df42['加班类型'] = Categorical(df42['加班类型'], categories=['工作日加班', '公休日加班', '法定节假日加班'], ordered=True)
# df42 = df42.sort_values(by=['序号', '加班类型'])
# df42 = df42.reset_index(drop=True)


# # 填充数据 加班费申报表
b4 = 5
for i4 in range(len(df42)):
    for j4 in range(len(df42.columns)):
        cell_value4 = df42.iat[i4, j4]
        ws_origin4.cell(i4 + b4, j4 + 1, value=cell_value4)
# # 合并单元格, 序号
# current_num4 = 5
# count4 = 0
# num4 = 1
# for k4 in range(len(df42)):
#     if k4 == 0:
#         continue
#     elif df42.iat[k4, 0] != df42.iat[k4 - 1, 0]:
#         num4 += count4
#         ws_origin4.merge_cells(f'A{current_num4}:A{current_num4 + count4}')
#         ws_origin4.merge_cells(f'B{current_num4}:B{current_num4 + count4}')
#         ws_origin4.merge_cells(f'C{current_num4}:C{current_num4 + count4}')
#         current_num4 = current_num4 + count4 + 1
#         count4 = 0
#     elif k4 == len(df42) - 1:
#         ws_origin4.merge_cells(f'A{current_num4}:A{current_num4 + count4 + 1}')
#         ws_origin4.merge_cells(f'B{current_num4}:B{current_num4 + count4 + 1}')
#         ws_origin4.merge_cells(f'C{current_num4}:C{current_num4 + count4 + 1}')
#     else:
#         count4 += 1

# # 合并审批人
# current_num41 = 5
# count41 = 0
# num41 = 1
# for k41 in range(len(df42)):
#     if k41 == 0:
#         continue
#     elif df42.iat[k41, 6] != df42.iat[k41 - 1, 6]:
#         num41 += count41
#         ws_origin4.merge_cells(f'G{current_num41}:G{current_num41 + count41}')
#         ws_origin4.merge_cells(f'H{current_num41}:H{current_num41 + count41}')
#         ws_origin4[f'G{current_num41}'] = ''
#         current_num41 = current_num41 + count41 + 1
#         count41 = 0
#     elif k41 == len(df42) - 1:
#         ws_origin4.merge_cells(f'G{current_num41}:G{current_num41 + count41 + 1}')
#         ws_origin4.merge_cells(f'H{current_num41}:H{current_num41 + count41 + 1}')
#         ws_origin4[f'G{current_num41}'] = ''
#     else:
#         count41 += 1


# ******************************5.加班审批表分公司-新************************************
df50 = handel_df_final(5, 'G', 3)

current_company5 = []
current_company5_list = []
if df50.empty:
    print("DataFrame is empty")
else:
    # 尝试删除列
    df50 = df50.drop(columns=['data_id_x', '本次加班获得调休时长', '画×', '备注', '报加班费还是给调休'])
    # 填充数据 加班审批表
    current_company5 = df50.iat[0, 13]
    current_company5_list = [df50.iat[0, 13]]

b5 = 5
num5 = 0
for i5 in range(len(df50)):
    for j5 in range(len(df50.columns)):
        cell_value = df50.iat[i5, j5]
        if df50.iat[i5, 13] != current_company5:
            num5 += 1
            current_company5 = df50.iat[i5, 13]
            current_company5_list.append(current_company5)
        ws_origin5.cell(i5 + b5 + num5, j5 + 1, value=cell_value)

# 啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊！
sum_list = []
count5 = 0
k52 = 0
sum_count = 5
for k5 in range(1, len(df50) + num5 + 1):
    # 合计那些行
    if ws_origin5[f'A{k5 + b5}'].value is None or ws_origin5[f'A{k5 + b5}'].value == '':
        sum_list.append(k5 + b5)
        # ws_origin5.merge_cells(f'A{k5 + b5}:K{k5 + b5}')
        ws_origin5.merge_cells(f'A{k5 + b5}:L{k5 + b5}')
        ws_origin5[f'A{k5 + b5}'] = f'{current_company5_list[k52]}小计'
        # ws_origin5[f'L{k5 + b5}'] = f'=SUM(L{sum_count}:L{k5 + b5 - 1})'
        ws_origin5[f'M{k5 + b5}'] = f'=SUM(M{sum_count}:M{k5 + b5 - 1})'
        ws_origin5[f'N{k5 + b5}'] = f'=SUM(N{sum_count}:N{k5 + b5 - 1})'
        ws_origin5[f'A{k5 + b5}'].fill = fill
        ws_origin5[f'L{k5 + b5}'].fill = fill
        ws_origin5[f'M{k5 + b5}'].fill = fill
        ws_origin5[f'N{k5 + b5}'].fill = fill
        ws_origin5[f'O{k5 + b5}'].fill = fill
        sum_count = k5 + b5 + 1
        k52 += 1
    # 正常行
    if ws_origin5[f'C{k5 + b5}'].value != ws_origin5[f'C{k5 + b5 - 1}'].value:
        if ws_origin5[f'C{k5 + b5 - 1 - count5}'].value is not None and ws_origin5[f'C{k5 + b5 - 1 - count5}'].value != '':
            ws_origin5[f'N{k5 + b5 - 1 - count5}'] = f'=SUM(M{k5 + b5 - 1 - count5}:M{k5 + b5 - 1})'
        ws_origin5.merge_cells(f'A{k5 + b5 - 1 - count5}:A{k5 + b5 - 1}')
        ws_origin5.merge_cells(f'B{k5 + b5 - 1 - count5}:B{k5 + b5 - 1}')
        ws_origin5.merge_cells(f'C{k5 + b5 - 1 - count5}:C{k5 + b5 - 1}')
        ws_origin5.merge_cells(f'N{k5 + b5 - 1 - count5}:N{k5 + b5 - 1}')
        count5 = 0
    else:
        count5 += 1

# 加班人数
ws_origin5['C250'] = len(df50['姓名'].unique())
# 法定节假日总时长
ws_origin5['E250'] = df50.loc[df50['加班类型'] == '法定节假日加班']['加班时长（小时）'].sum()
# 公休日总时长
ws_origin5['H250'] = df50.loc[df50['加班类型'] == '公休日加班']['加班时长（小时）'].sum()
# 工作日总时长
ws_origin5['K250'] = df50.loc[df50['加班类型'] == '工作日加班']['加班时长（小时）'].sum()
result11 = "+".join([f"M{num}" for num in sum_list])
result22 = "+".join([f"N{num}" for num in sum_list])
ws_origin5['M250'] = "=" + result11
ws_origin5['N250'] = "=" + result22



# ******************************6.领导画×表************************************
df60 = handel_df_final(5, 'G', 3)
current_company6 = []
current_company6_list = []
if df60.empty:
    print("DataFrame is empty, cannot access values.")
else:
    # DataFrame 不为空，可以安全地访问值
    # 填充数据 加班审批表
    current_company6 = df60.iat[0, 13]
    current_company6_list = [df60.iat[0, 13]]
b6 = 5
num6 = 0
for i6 in range(len(df60)):
    for j6 in range(len(df60.columns)):
        cell_value = df60.iat[i6, j6]
        if df60.iat[i6, 13] != current_company6:
            num6 += 1
            current_company6 = df60.iat[i6, 13]
            current_company6_list.append(current_company6)
        ws_origin6.cell(i6 + b6 + num6, j6 + 1, value=cell_value)

# 啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊！
sum_list = []
count6 = 0
k62 = 0
sum_count = 5
for k6 in range(1, len(df60) + num6 + 1):
    # 合计那些行
    if ws_origin6[f'A{k6 + b6}'].value is None or ws_origin6[f'A{k6 + b6}'].value == '':
        sum_list.append(k6 + b6)
        ws_origin6.merge_cells(f'A{k6 + b6}:K{k6 + b6}')
        ws_origin6[f'A{k6 + b6}'] = f'{current_company6_list[k62]}小计'
        ws_origin6[f'L{k6 + b6}'] = f'=SUM(L{sum_count}:L{k6 + b6 - 1})'
        ws_origin6[f'M{k6 + b6}'] = f'=SUM(M{sum_count}:M{k6 + b6 - 1})'
        ws_origin6[f'N{k6 + b6}'] = f'=SUM(N{sum_count}:N{k6 + b6 - 1})'
        ws_origin6[f'A{k6 + b6}'].fill = fill
        ws_origin6[f'L{k6 + b6}'].fill = fill
        ws_origin6[f'M{k6 + b6}'].fill = fill
        ws_origin6[f'N{k6 + b6}'].fill = fill
        ws_origin6[f'O{k6 + b6}'].fill = fill
        sum_count = k6 + b6 + 1
        k62 += 1
    # 正常行
    if ws_origin6[f'C{k6 + b6}'].value != ws_origin6[f'C{k6 + b6 - 1}'].value:
        if ws_origin6[f'C{k6 + b6 - 1 - count6}'].value is not None and ws_origin6[f'C{k6 + b6 - 1 - count6}'].value != '':
            ws_origin6[f'N{k6 + b6 - 1 - count6}'] = f'=SUM(M{k6 + b6 - 1 - count6}:M{k6 + b6 - 1})'
        ws_origin6.merge_cells(f'A{k6 + b6 - 1 - count6}:A{k6 + b6 - 1}')
        ws_origin6.merge_cells(f'B{k6 + b6 - 1 - count6}:B{k6 + b6 - 1}')
        ws_origin6.merge_cells(f'C{k6 + b6 - 1 - count6}:C{k6 + b6 - 1}')
        ws_origin6.merge_cells(f'N{k6 + b6 - 1 - count6}:N{k6 + b6 - 1}')
        count6 = 0
    else:
        count6 += 1

# 表头
ws_origin6['O4'] = '画×'
ws_origin6['P4'] = '报加班费还是给调休'
ws_origin6['Q4'] = '本次加班获得调休时长'
ws_origin6['R4'] = 'data_id'
# 加班人数
ws_origin6['C250'] = len(df60['姓名'].unique())
# 法定节假日总时长
ws_origin6['E250'] = df60.loc[df60['加班类型'] == '法定节假日加班']['加班时长（小时）'].sum()
# 公休日总时长
ws_origin6['G250'] = df60.loc[df60['加班类型'] == '公休日加班']['加班时长（小时）'].sum()
# 工作日总时长
ws_origin6['J250'] = df60.loc[df60['加班类型'] == '工作日加班']['加班时长（小时）'].sum()
result11 = "+".join([f"L{num}" for num in sum_list])
result22 = "+".join([f"M{num}" for num in sum_list])
ws_origin6['L250'] = "=" + result11
ws_origin6['M250'] = "=" + result22
df_unique = df14.drop_duplicates(subset='姓名', keep='first')
# df_unique.to_excel('/home/sf107/桌面/33333333333.xlsx', index=False)
filtered_rows = df_unique[df_unique['人员类别'].isin(['党务专职（直聘）', '党务专职（参聘）', '社区专职（直聘）', '社区专职（参聘）', '社区专职（未过渡）'])]
count_p = len(filtered_rows)
# 单位聘用人员（含直聘、参聘、党务）总人数
ws_origin4['D100'] = count_p
# 实际加班人数：
ws_origin4['D101'] = len(df60['姓名'].unique())


# **************************7.加班费申报表（季度）*********************************
# # 填充数据 加班费申报表（季度）
b4 = 5
for i4 in range(len(df42)):
    for j4 in range(len(df42.columns)):
        cell_value4 = df42.iat[i4, j4]
        ws_origin7.cell(i4 + b4, j4 + 1, value=cell_value4)




ws_origin7['D4'] = f"第{current_quarter}季度加班费\n总金额（元）"
ws_origin7['A111'] = f"{current_year}年{month1}月"
ws_origin7['C111'] = f"{current_year}年{month2}月"
ws_origin7['E111'] = f"{current_year}年{month3}月"


def get_people_count(month):
    # 筛选出'名单生效年月_年月'列年份为current_year且月份为month1的所有行
    filtered_df1 = df14.loc[(df14['名单生效年月_年月'].dt.year == current_year) & (df14['名单生效年月_年月'].dt.month == month)]
    # 进一步筛选'人员类别'列在指定列表中的行
    filtered_rows1 = filtered_df1[filtered_df1['人员类别'].isin(['党务专职（直聘）', '党务专职（参聘）', '社区专职（直聘）', '社区专职（参聘）', '社区专职（未过渡）'])]
    # 计算进一步筛选后的行数
    return len(filtered_rows1)


ws_origin7['B112'] = get_people_count(month1)
ws_origin7['D112'] = get_people_count(month2)
ws_origin7['F112'] = get_people_count(month3)
ws_origin7['B113'] = len(month1_df['姓名'].unique())
ws_origin7['D113'] = len(month2_df['姓名'].unique())
ws_origin7['F113'] = len(month3_df['姓名'].unique())
ws_origin7['B114'] = month1_df['加班费金额'].sum()
ws_origin7['D114'] = month2_df['加班费金额'].sum()
ws_origin7['F114'] = month3_df['加班费金额'].sum()

# 表格里面的标题
# ws_origin.cell(1, 1, value=f"民治街道上芬社区{current_year}年{current_month}月加班审批表")
ws_origin2.cell(2, 1, value=f"民治街道办事处（上芬社区）{current_year}年{current_month}月加班日志汇总表\n（注：已调休部分以及确认调休部分无需填报，各部门保管好相关材料留底。）")
# ws_origin3.cell(1, 1, value=f"民治街道上芬社区{current_year}年{current_month}月单日加班超4小时申请汇总表")
ws_origin4.cell(2, 1, value=f"民治街道办事处（上芬社区）{current_year}年{current_month}月加班费申报表")
ws_origin5.cell(2, 1, value=f"民治街道办事处（上芬社区）{current_year}年{current_month}月加班费审批表")
ws_origin6.cell(1, 1, value=f"民治街道上芬社区{current_year}年{current_month}月加班费审批表")
ws_origin7.cell(2, 1, value=f"民治街道办事处（上芬社区）{current_year % 100}年第{current_quarter}季度加班费申报表")

# 将工作簿保存为 Excel 文件
# wb_origin.save(f"{path}处理结果/原来的加班费审批表{current_month}月.xlsx")  # ！！！
wb_origin2.save(f"{path}处理结果/{current_year}{current_month}附件3：加班日志汇总表.xlsx")  # ！！！
# wb_origin3.save(f"{path}处理结果/单日加班超4小时申请汇总表.xlsx")  # ！！！
wb_origin4.save(f"{path}处理结果/{current_year}{current_month}附件1：加班费申报表.xlsx")  # ！！！
wb_origin5.save(f"{path}处理结果/{current_year}{current_month}附件2：加班费审批表 .xlsx")  # ！！！
wb_origin6.save(f"{path}处理结果/{current_month}月领导画×表.xlsx")  # ！！！
wb_origin7.save(f"{path}处理结果/附件1：加班费申报表（第{current_quarter}季度）.xlsx")  # ！！！

# subprocess.run(['python', '处理调休表.py'])
# result = sum(i*i for i in range(0,101))
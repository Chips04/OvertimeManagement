import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import fnmatch
from datetime import timedelta
import json
import subprocess

# 读取txt文件内容
with open('variables.txt', 'r') as file:
    content = file.read().strip()

# 尝试将内容转换为字典，注意这里假设了txt文件的内容是合法的JSON格式
try:
    data = json.loads(content)  # 将字符串转换为字典
except json.JSONDecodeError as e:
    print(f"无法解析txt文件内容为JSON: {e}")
    exit(1)

# 从字典中提取变量
path = data['path']
month = data['month']
year = data['year']


# 调休表需要导出所有日期加班和所有请休假
df00 = pd.read_excel(path + '工作日周末法定节假日表_20240515175839.xlsx')
# 将DataFrame中的日期列转换为datetime格式
df00['日期1'] = pd.to_datetime(df00['日期'])
# 只保留工作日请假的行
df0 = df00[df00['类型'] == '工作日加班']

# 获取数据源dataframe
df1 = pd.DataFrame()
# 指定目录和通配符
directory = path
directory2 = path + '调休处理数据源/'
pattern = '加班申请*.xlsx'
# 获取匹配的文件列表
matches = []
for filename in os.listdir(directory2):
    if fnmatch.fnmatch(filename, pattern):
        matches.append(os.path.join(directory2, filename))

# 打开第一个匹配的文件并读取内容
if matches:
    df1 = pd.read_excel(matches[0])
else:
    print('No matching files found.')
if month < 10:
    month_str = "2024.0" + str(month)
else:
    month_str = "2024." + str(month)
# 获得年月
max_value = df1['加班日期'].max()
current_year = max_value.year
current_month = max_value.month
wb_origin7 = openpyxl.load_workbook(path + '模板/补休情况登记表模板.xlsx')  # ！！！
ws_origin7 = wb_origin7.active
ws_origin7.cell(1, 1, value=f"民治街道上芬社区{current_year}年{current_month}月补休情况登记表")
# 大表
wb_origin9 = openpyxl.load_workbook(path + '模板/补休情况登记表模板大表模板.xlsx')  # ！！！
ws_origin9 = wb_origin9.active
ws_origin9.cell(1, 1, value=f"民治街道上芬社区{current_year}年{current_month}月补休情况登记表")


# 获取数据源dataframe
df2 = pd.DataFrame()

pattern2 = '社区工作人员花名册*.xlsx'
# 获取匹配的文件列表
matches2 = []
for filename2 in os.listdir(directory):
    if fnmatch.fnmatch(filename2, pattern2):
        matches2.append(os.path.join(directory, filename2))

# 打开第一个匹配的文件并读取内容
if matches2:
    df2 = pd.read_excel(matches2[0])
else:
    print('No matching files found2.')


# 获取数据源dataframe
df3 = pd.DataFrame()
# 指定目录和通配符
pattern3 = '人员管理-请休假*.xlsx'
# 获取匹配的文件列表
matches3 = []
for filename3 in os.listdir(directory):
    if fnmatch.fnmatch(filename3, pattern3):
        matches3.append(os.path.join(directory, filename3))

# 打开第一个匹配的文件并读取内容
if matches3:
    df3 = pd.read_excel(matches3[0])
else:
    print('No matching files found.')


# 获得年月
max_value = df1['加班日期'].max()
current_year = max_value.year
current_month = max_value.month

# 只保留调休的行
df1 = df1[df1['报加班费还是给调休'].str.contains('调休')]


# 删除不同意的行
if '初核意见' in df1.columns:
    df1 = df1[df1['初核意见'] != '不同意']
    df1 = df1[df1['复核意见'] != '不同意']
    df1 = df1[df1['初初核意见'] != '不同意']


df2 = df2.rename(columns={'工作人员姓名': '姓名'})

# 合并df1和df2
result = pd.merge(df1, df2, on='姓名', how='left')

# 排序
result = result.sort_values(by=['编号', '加班开始时间'], ascending=[True, True])

result = result.reset_index(drop=True)

# 结果的表格
columns2 = ['序号', '姓名', '加班日期\n（附时间段）', '补休日期\n（附时间段）', '加班时长', '补休时长', '剩余时长', '备注', '1', '2']


def convert_time(time_list):
    saved_date = None
    temp_date = None
    str1 = ''
    last_date = ''
    for i, item in enumerate(time_list):
        if str1 == '':
            last_date = item.strftime("%Y.%m.%d")
            str1 = last_date
            temp_date = item
            saved_date = item
        elif item - temp_date == timedelta(hours=0.5):
            temp_date = item
        else:
            str1 += f"\n（{saved_date.strftime('%H:%M')}-{(temp_date + timedelta(hours=0.5)).strftime('%H:%M')}）"
            if item.strftime("%Y.%m.%d") != last_date:
                last_date = item.strftime("%Y.%m.%d")
                str1 += '\n' + last_date
            saved_date = item
            temp_date = item
        if i == len(time_list) - 1:
            str1 += f"\n（{saved_date.strftime('%H:%M')}-{(temp_date + timedelta(hours=0.5)).strftime('%H:%M')}）"
    time_list_len = len(time_list) / 2
    return str1, time_list_len


# 请休假表格处理
# 只保留调休的行
df3 = df3[df3['休假类别'] == '调休']
# 只保留做表格之前的，不要之后的请休假
# 筛选出请假开始日期在4月以前（包括4月）的行
condition_date = df3['请假开始日期'].dt.month <= month
# 筛选出请假开始时间在4月以前（包括4月）的行
# 注意：这里假设leave_start_time列包含日期和时间信息
condition_time = df3['调休开始时间（不足半天）'].dt.month <= month
# 结合两个条件进行筛选
df3 = df3[(condition_date) | (condition_time)]

df3 = df3.sort_values(by=['申请人员姓名', '请假开始日期', '请假开始时段'])
df3 = df3.reset_index(drop=True)

df3.loc[df3['请假开始时段（网格）'].notnull(), '请假开始时段'] = df3['请假开始时段（网格）']
df3.loc[df3['请假结束时段（网格）'].notnull(), '请假结束时段'] = df3['请假结束时段（网格）']
df3['请假开始时段'] = df3['请假开始时段'].replace({'前半天': '上午', '后半天': '下午'})
df3['请假结束时段'] = df3['请假结束时段'].replace({'前半天': '上午', '后半天': '下午'})

df3['调休时间list'] = pd.Series([[] for _ in range(len(df3))], dtype=object)


# 当天是什么班
def get_mr(date):
    return df00.loc[df00['日期'] == date.strftime('%Y-%m-%d'), '网格班次']


for index3, row3 in df3.iterrows():
    if row3['是否【不足半天】'] == '是':
        num_intervals4 = int((row3['调休结束时间（不足半天）'] - row3['调休开始时间（不足半天）']).total_seconds() / 1800)
        row3['调休时间list'] += [row3['调休开始时间（不足半天）'] + timedelta(minutes=i * 30) for i in range(num_intervals4)]
    elif row3['请假开始时段'] == row3['请假结束时段']:
        row3['调休时间list'] = []
        if row3['请假开始时段'] == '上午':  # 上午上午即开始整天加结束上午半天
            i22 = 0
            for i2 in range((row3['请假结束日期'] - row3['请假开始日期']).days):
                matching_row0 = get_mr(row3['请假开始日期'] + timedelta(days=i2))
                if row3['计算人员类别'] == '晚班人' and not matching_row0.empty and '下午晚上班' == matching_row0.values[0]:
                    list21 = [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(days=i2) + timedelta(minutes=i * 30) for i in range(7)]
                    list22 = [row3['请假开始日期'].replace(hour=19, minute=30, second=0, microsecond=0) + timedelta(days=i2) + timedelta(minutes=i * 30) for i in range(6)]
                else:
                    list21 = [row3['请假开始日期'].replace(hour=9, minute=0, second=0, microsecond=0) + timedelta(days=i2) + timedelta(minutes=i * 30) for i in range(6)]
                    list22 = [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(days=i2) + timedelta(minutes=i * 30) for i in range(8)]
                row3['调休时间list'] = row3['调休时间list'] + list21 + list22
                i22 = i2
            # 创建时间点列表，不包括最后一个时间点（即end_time）
            matching_row = get_mr(row3['请假开始日期'] + timedelta(days=i22 + 1))
            if row3['计算人员类别'] == '晚班人' and not matching_row.empty and '下午晚上班' == matching_row.values[0]:
                row3['调休时间list'] = row3['调休时间list'] + [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(days=i22 + 1) + timedelta(minutes=i * 30) for i in range(7)]
            else:
                row3['调休时间list'] = row3['调休时间list'] + [row3['请假开始日期'].replace(hour=9, minute=0, second=0, microsecond=0) + timedelta(days=i22 + 1) + timedelta(minutes=i * 30) for i in range(6)]
            # print(111111111)
        else:  # 下午下午即开始下午半天加结束整天
            matching_row = get_mr(row3['请假开始日期'])
            if row3['计算人员类别'] == '晚班人' and not matching_row.empty and '下午晚上班' == matching_row.values[0]:
                row3['调休时间list'] = [row3['请假开始日期'].replace(hour=19, minute=30, second=0, microsecond=0) + timedelta(minutes=i * 30) for i in range(6)]
            else:
                row3['调休时间list'] = [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(minutes=i * 30) for i in range(8)]
            for i1 in range((row3['请假结束日期'] - row3['请假开始日期']).days):
                matching_row0 = get_mr(row3['请假开始日期'] + timedelta(days=i1 + 1))
                if row3['计算人员类别'] == '晚班人' and not matching_row0.empty and '下午晚上班' == matching_row0.values[0]:
                    list11 = [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(days=i1 + 1) + timedelta(minutes=i * 30) for i in range(7)]
                    list12 = [row3['请假开始日期'].replace(hour=19, minute=30, second=0, microsecond=0) + timedelta(days=i1 + 1) + timedelta(minutes=i * 30) for i in range(6)]
                else:
                    list11 = [row3['请假开始日期'].replace(hour=9, minute=0, second=0, microsecond=0) + timedelta(days=i1 + 1) + timedelta(minutes=i * 30) for i in range(6)]
                    list12 = [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(days=i1 + 1) + timedelta(minutes=i * 30) for i in range(8)]
                row3['调休时间list'] = row3['调休时间list'] + list11 + list12
            # 创建时间点列表，不包括最后一个时间点（即end_time）
            # print(22222222222)
    elif row3['请假开始时段'] == '上午':  # 开始是上午，结束是下午，即日期差整天
        row3['调休时间list'] = []
        for i4 in range((row3['请假结束日期'] - row3['请假开始日期']).days + 1):
            matching_row0 = get_mr(row3['请假开始日期'] + timedelta(days=i4))
            if row3['计算人员类别'] == '晚班人' and not matching_row0.empty and '下午晚上班' == matching_row0.values[0]:
                list41 = [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(days=i4) + timedelta(minutes=i * 30) for i in range(7)]
                list42 = [row3['请假开始日期'].replace(hour=19, minute=30, second=0, microsecond=0) + timedelta(days=i4) + timedelta(minutes=i * 30) for i in range(6)]
            else:
                list41 = [row3['请假开始日期'].replace(hour=9, minute=0, second=0, microsecond=0) + timedelta(days=i4) + timedelta(minutes=i * 30) for i in range(6)]
                list42 = [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(days=i4) + timedelta(minutes=i * 30) for i in range(8)]
            row3['调休时间list'] = row3['调休时间list'] + list41 + list42
        # print(333333333)
    else:  # 日期不同，开始是下午，结束是上午，即开始下午加中间日期加结束上午
        matching_row = get_mr(row3['请假开始日期'])
        if row3['计算人员类别'] == '晚班人' and not matching_row.empty and '下午晚上班' == matching_row.values[0]:
            row3['调休时间list'] = [row3['请假开始日期'].replace(hour=19, minute=30, second=0, microsecond=0) + timedelta(minutes=i * 30) for i in range(6)]
        else:
            row3['调休时间list'] = [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(minutes=i * 30) for i in range(8)]
        i55 = 0
        for i5 in range((row3['请假结束日期'] - row3['请假开始日期']).days - 1):
            matching_row0 = get_mr(row3['请假开始日期'] + timedelta(days=i5 + 1))
            if row3['计算人员类别'] == '晚班人' and not matching_row0.empty and '下午晚上班' == matching_row0.values[0]:
                list51 = [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(days=i5 + 1) + timedelta(minutes=i * 30) for i in range(7)]
                list52 = [row3['请假开始日期'].replace(hour=19, minute=30, second=0, microsecond=0) + timedelta(days=i5 + 1) + timedelta(minutes=i * 30) for i in range(6)]
            else:
                list51 = [row3['请假开始日期'].replace(hour=9, minute=0, second=0, microsecond=0) + timedelta(days=i5 + 1) + timedelta(minutes=i * 30) for i in range(6)]
                list52 = [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(days=i5 + 1) + timedelta(minutes=i * 30) for i in range(8)]
            row3['调休时间list'] = row3['调休时间list'] + list51 + list52
            i55 = i5
        matching_row1 = get_mr(row3['请假开始日期'] + timedelta(days=i55 + 2))
        if row3['计算人员类别'] == '晚班人' and not matching_row1.empty and '下午晚上班' == matching_row1.values[0]:
            row3['调休时间list'] = row3['调休时间list'] + [row3['请假开始日期'].replace(hour=14, minute=0, second=0, microsecond=0) + timedelta(days=i55 + 2) + timedelta(minutes=i * 30) for i in range(7)]
        else:
            row3['调休时间list'] = row3['调休时间list'] + [row3['请假开始日期'].replace(hour=9, minute=0, second=0, microsecond=0) + timedelta(days=i55 + 2) + timedelta(minutes=i * 30) for i in range(6)]

        # print(444444444)
    # 删除周末和法定节假日
    # 找出不在DataFrame日期列中的日期，并创建新的列表
    # filtered_date_list = [date for date in row3['调休时间list'] if date.strftime('%Y-%m-%d %H:%M:%S') not in df0['日期1']]
    # filtered_date_list = [date for date in row3['调休时间list'] if date.strftime('%Y-%m-%d %H:%M:%S') not in df0['日期1']]
    # print('======================')
    # filtered_date_list = [date for date in row3['调休时间list'] if date.replace(hour=0, minute=0, second=0, microsecond=0) not in df0['日期1']]

    dates_set = set(df0['日期1'].dt.date)

    filtered_timestamp_list = [
        timestamp for timestamp in row3['调休时间list']
        if timestamp.date() in dates_set
    ]
    df3.at[index3, '调休时间list'] = filtered_timestamp_list
    # print(row3['调休时间list'])
    # print(len(row3['调休时间list']))


# 使用groupby和agg函数合并行
df4 = df3.groupby('申请人员姓名', as_index=False).agg({
    '调休时间list': lambda x: [item for sublist in x for item in sublist]  # 累加列表
})
df4['调休时间list'] = df4['调休时间list'].apply(sorted)
df4['list_length'] = df4['调休时间list'].apply(len)
# print(df4)


df7 = pd.DataFrame(columns=columns2)
row_index = 0
row_name = ''
result = result.reset_index()
for index_final, row_final in result.iterrows():
    # 将字符串转换为datetime对象
    start_datetime = row_final['加班开始时间']
    end_datetime = row_final['加班结束时间']
    # 格式化输出
    start_formatted = start_datetime.strftime("%Y.%m.%d")
    start_formatted2 = start_datetime.strftime("%H:%M")
    end_formatted = end_datetime.strftime("%H:%M")

    if not pd.isna(row_final['加班开始时间2']):
        start_datetime2 = row_final['加班开始时间2']
        start_formatted4 = start_datetime2.strftime("%H:%M")
    if not pd.isna(row_final['加班结束时间2']):
        end_datetime2 = row_final['加班结束时间2']
        end_formatted2 = end_datetime2.strftime("%H:%M")
    if not pd.isna(row_final['加班开始时间3']):
        start_datetime3 = row_final['加班开始时间3']
        start_formatted6 = start_datetime3.strftime("%H:%M")
    if not pd.isna(row_final['加班结束时间3']):
        end_datetime3 = row_final['加班结束时间3']
        end_formatted3 = end_datetime3.strftime("%H:%M")

    time_points1 = []
    time_points2 = []
    time_points3 = []
    # 计算从开始时间到结束时间之间的完整30分钟间隔数量
    num_intervals1 = int((row_final['加班结束时间'] - row_final['加班开始时间']).total_seconds() / 1800)
    # 创建时间点列表，不包括最后一个时间点（即end_time）
    time_points1 = [row_final['加班开始时间'] + timedelta(minutes=i * 30) for i in range(num_intervals1)]
    if row_final['你这一天分了几段时间'] == '一段':
        # 构造最终格式的字符串
        time1 = f'{start_formatted}\n（{start_formatted2}-{end_formatted}）'
    elif row_final['你这一天分了几段时间'] == '两段':
        # 构造最终格式的字符串
        time1 = f'{start_formatted}\n（{start_formatted2}-{end_formatted}）\n（{start_formatted4}-{end_formatted2}）'
        # 计算从开始时间到结束时间之间的完整30分钟间隔数量
        num_intervals2 = int((row_final['加班结束时间2'] - row_final['加班开始时间2']).total_seconds() / 1800)
        # 创建时间点列表，不包括最后一个时间点（即end_time）
        time_points2 = [row_final['加班开始时间2'] + timedelta(minutes=i * 30) for i in range(num_intervals2)]
    else:
        # 构造最终格式的字符串
        time1 = f'{start_formatted}\n（{start_formatted2}-{end_formatted}）\n（{start_formatted4}-{end_formatted2}）\n（{start_formatted6}-{end_formatted3}）'
        # 计算从开始时间到结束时间之间的完整30分钟间隔数量
        num_intervals2 = int((row_final['加班结束时间2'] - row_final['加班开始时间2']).total_seconds() / 1800)
        # 创建时间点列表，不包括最后一个时间点（即end_time）
        time_points2 = [row_final['加班开始时间2'] + timedelta(minutes=i * 30) for i in range(num_intervals2)]
        # 计算从开始时间到结束时间之间的完整30分钟间隔数量
        num_intervals3 = int((row_final['加班结束时间3'] - row_final['加班开始时间3']).total_seconds() / 1800)
        # 创建时间点列表，不包括最后一个时间点（即end_time）
        time_points3 = [row_final['加班开始时间3'] + timedelta(minutes=i * 30) for i in range(num_intervals3)]

    time_points = time_points1 + time_points2 + time_points3

    # 调休日期
    time2 = ''
    row_to_modify = pd.Series()
    flag = 0
    time_list_len = 0
    # 检查是否有姓名为“张三”的行
    if row_final['姓名'] in df4['申请人员姓名'].values:
        flag = 1
        # 定位到姓名为“张三”的行
        row_to_modify = df4.loc[df4['申请人员姓名'] == row_final['姓名']].iloc[0]
        time2, time_list_len = convert_time(df4.loc[df4['申请人员姓名'] == row_final['姓名']].iloc[0]['调休时间list'][:len(time_points)])
    else:
        # print(f"没有找到姓名为{row_final['姓名']}的行。")
        pass
    newRow1 = pd.Series(
        [
            '',
            row_final['姓名'],
            time1,
            time2,
            row_final["加班计算小时数"],
            time_list_len,
            row_final["加班计算小时数"] - time_list_len,
            '',
            '',
            ''
        ],
        index=columns2
    )
    if flag == 1:
        if len(row_to_modify['调休时间list']) > 0 and len(time_points) > 0:
            if row_to_modify['调休时间list'][len(row_to_modify['调休时间list']) - 1] < time_points[len(time_points) - 1]:
                # 请的假期用了当天以后的调休
                print(f"{row_final['姓名']}的时间不对哦！")
        later_list = row_to_modify['调休时间list'][len(time_points):]
        indices = df4[df4['申请人员姓名'] == row_final['姓名']].index[0]
        # 这一行代码费我好大劲，纪念！
        df4.iat[indices, 1] = later_list

    df7 = df7.append(newRow1, ignore_index=True)

# 数组要是有不为空的 那就是假请多了↓
print(df4)

# ******************************7.补休情况登记表************************************
df7['序号'] = range(1, len(df7) + 1)
df7['1'] = df7['姓名'].map(df2.set_index('姓名')['所属部门'])
print(df2.columns)
print(df7.columns)
# # 填充数据 大表
b9 = 4
for i9 in range(len(df7)):
    for j9 in range(len(df7.columns)):
        cell_value4 = df7.iat[i9, j9]
        ws_origin9.cell(i9 + b9, j9 + 1, value=cell_value4)
# 输不输出大表
# wb_origin9.save(f"{path}处理结果/附件6：补休情况登记表-大表-全.xlsx")  # ！！！
unique_names = list(dict.fromkeys(df7['姓名']))
print(unique_names)
for name in unique_names:
    copied_worksheet = wb_origin7.copy_worksheet(wb_origin7['模板'])
    copied_worksheet.title = name
    # 填充数据 补休情况登记表
    b7 = 4
    df_new = df7[df7['姓名'] == name]
    for i7 in range(len(df_new)):
        copied_worksheet.cell(i7 + b7, 1, value=i7 + 1)
        for j7 in range(len(df_new.columns)):
            cell_value = df_new.iat[i7, j7]
            if cell_value != '':
                copied_worksheet.cell(i7 + b7, j7 + 1, value=cell_value)

wb_origin7.remove(wb_origin7['模板'])
# 将工作簿保存为 Excel 文件

wb_origin7.save(f"{path}处理结果/附件6：补休情况登记表-全.xlsx")  # ！！！
wb_origin8 = openpyxl.load_workbook(path + '处理结果/附件6：补休情况登记表-全.xlsx')  # ！！！


# 遍历所有工作表
for sheet in wb_origin8.worksheets:
    # print(f"正在处理工作表: {sheet.title}")
    flag = 0
    # 创建一个列表来存储需要保留的行号
    rows_to_keep = []

    # 遍历工作表的每一行（从第四行开始）
    for row_num, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        # 遍历行中的每个单元格
        for cell in row:
            # 检查单元格内容是否包含'month_str'
            if month_str in str(cell):
                # 如果包含，则将该行号添加到保留列表中
                rows_to_keep.append(row_num)
                flag = 1
                break  # 既然已经找到了匹配项，就可以跳出内层循环了
    if flag == 0:
        wb_origin8.remove(wb_origin8[sheet.title])
    else:
        # 获取工作表的最后一行的行号
        last_row = sheet.max_row

        # 创建一个列表来存储需要删除的行号
        rows_to_delete = []

        # 遍历工作表的每一行（从第四行开始）
        for row_num, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            # 如果该行不在保留列表中且不是最后两行，则添加到删除列表中
            if row_num not in rows_to_keep and row_num != last_row - 1 and row_num != last_row:
                rows_to_delete.append(row_num)

        # 从后向前删除行，以避免影响行号
        for row_num in reversed(rows_to_delete):
            sheet.delete_rows(row_num)
        last_row2 = sheet.max_row
        # print(sheet['A'+str(last_row2 - 1)].value)
        sheet['A'+str(last_row2 - 1)].alignment = openpyxl.styles.Alignment(wrap_text=False)
        for i in range(last_row2 - 1, 50):
            sheet.row_dimensions[i].height = 14.25
        for j in range(4,last_row2 - 1):
            # print(j - 3)
            # print(sheet['D'+str(j)].value)
            cell_value = sheet['D' + str(j)].value
            if month_str not in str(cell_value):
                # print(111111111111111)
                # print(cell_value)
                sheet['D' + str(j)] = None
            sheet['A'+str(j)] = j - 3
        sheet.cell(1, 1, value=f"民治街道上芬社区{current_year}年{month}月补休情况登记表")

# 大表
# 遍历所有工作表
# print(f"正在处理工作表: {sheet.title}")
flag = 0
# 创建一个列表来存储需要保留的行号
rows_to_keep = []

# 遍历工作表的每一行（从第四行开始）
sheet2 = wb_origin9.active
for row_num, row in enumerate(sheet2.iter_rows(min_row=4, values_only=True), start=4):
    # 遍历行中的每个单元格
    for cell in row:
        # 检查单元格内容是否包含'month_str'
        if month_str in str(cell):
            # 如果包含，则将该行号添加到保留列表中
            rows_to_keep.append(row_num)
            flag = 1
            break  # 既然已经找到了匹配项，就可以跳出内层循环了

# 获取工作表的最后一行的行号
last_row = sheet2.max_row

# 创建一个列表来存储需要删除的行号
rows_to_delete2 = []

# 遍历工作表的每一行（从第四行开始）
for row_num, row in enumerate(sheet2.iter_rows(min_row=4, values_only=True), start=4):
    # 如果该行不在保留列表中且不是最后两行，则添加到删除列表中
    if row_num not in rows_to_keep and row_num != last_row - 1 and row_num != last_row:
        rows_to_delete2.append(row_num)

# 从后向前删除行，以避免影响行号
for row_num in reversed(rows_to_delete2):
    sheet2.delete_rows(row_num)
last_row2 = sheet2.max_row
# print(sheet2['A'+str(last_row2 - 1)].value)
sheet2['A'+str(last_row2 - 1)].alignment = openpyxl.styles.Alignment(wrap_text=False)
for i in range(last_row2 - 1, 50):
    sheet2.row_dimensions[i].height = 14.25
for j in range(4,last_row2 - 1):
    # print(j - 3)
    # print(sheet2['D'+str(j)].value)
    cell_value = sheet2['D' + str(j)].value
    if month_str not in str(cell_value):
        # print(111111111111111)
        # print(cell_value)
        sheet2['D' + str(j)] = None
    sheet2['A'+str(j)] = j - 3
print(month)
sheet2.cell(1, 1, value=f"民治街道上芬社区{current_year}年{month}月补休情况登记表")

# 初始化变量
start_row = 4
current_name = None

# 从第4行开始遍历B列
for row in range(start_row, sheet2.max_row + 1):
    # 获取当前行的姓名
    name = sheet2.cell(row=row, column=2).value

    # 如果当前姓名与上一个姓名相同，则继续遍历
    if name == current_name:
        continue

    # 如果当前姓名与上一个姓名不同，则合并上一个姓名对应的单元格
    if current_name is not None:
        # 计算要合并的单元格范围
        end_row = row - 1
        # 假设你想要合并从第二列（B列）的 start_row 到 end_row 的所有单元格
        sheet2.merge_cells(range_string=f'B{start_row}:B{end_row}')
        sheet2.merge_cells(range_string=f'H{start_row}:H{end_row}')
        # sheet2.merge_cells(start_row=start_row, end_row=end_row, start_col=2, end_col=2)

    # 更新起始行和当前姓名
    start_row = row
    current_name = name

# 合并最后一个姓名对应的单元格（如果有的话）
if current_name is not None:
    end_row = sheet2.max_row
    # 假设你想要合并从第二列（B列）的 start_row 到 end_row 的所有单元格
    sheet2.merge_cells(range_string=f'B{start_row}:B{end_row}')
    sheet2.merge_cells(range_string=f'H{start_row}:H{end_row}')
    # sheet2.merge_cells(start_row=start_row, end_row=end_row, start_col=2, end_col=2)

# 获取工作表的最后一行
max_row = sheet2.max_row
print(max_row)

# 从第 4 行遍历到倒数第 3 行
for row in range(4, max_row - 1):  # 减 2 是因为我们要到倒数第 3 行
    # 获取 D 列的单元格对象
    cell = sheet2.cell(row=row, column=4)  # 列 D 是第 4 列

    # 检查单元格是否为空
    if not cell.value:
        # 如果为空，则填充 "未补休"
        cell.value = "未补休"

# 保存工作簿
# wb_origin8.save('处理后的文件路径.xlsx')

wb_origin8.save(f"{path}处理结果/附件6：补休情况登记表-{month}月.xlsx")  # ！！！
wb_origin9.save(f"{path}处理结果/附件6：补休情况登记表-{month}月-大表.xlsx")  # ！！！

# subprocess.run(['python', '处理年假表.py'])
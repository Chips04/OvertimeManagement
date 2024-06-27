import pandas as pd
from datetime import datetime
from datetime import timedelta
from functions import *
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

filename5 = '/home/sf107/桌面/值班安排表处理/工作日周末法定节假日表_20240401151558.xlsx'
filename = '/home/sf107/桌面/值班安排表处理/上芬社区应急值班安排表2024年7月份  .xlsx'
output_name = '/home/sf107/桌面/值班安排表处理/值班安排表输出表格.xlsx'
output_name2 = '/home/sf107/桌面/值班安排表处理/值班安排表有值班领导输出表格.xlsx'
df = pd.read_excel(filename, skiprows=1)
df5 = pd.read_excel(filename5)
df11 = df5.copy()
# 只保留节假日值班的行
df5 = df5[df5['值班（节假日）'] == '节假日值班']
leader_arr = ['徐军', '田普洲', '何欢', '肖志强', '李超结', '郭亚飞', '龙倩琪', '刘波哲', '陈学荣', '王雄', '苏伟如', '戴南真', '马贵鑫', '詹国香', '罗有杰', '万成兵']
repeated_list = [x for i in range(len(df5)) for x in leader_arr]

df5['值班领导'] = repeated_list[:len(df5)]
year = filename.split("年")[0].split("值班安排表")[-1].strip()

# 获取最后一行第一列的值
last_row_first_column_value = df.iloc[-1]['序号']

# 检查该值是否为数字字符串
if not last_row_first_column_value.isdigit():
    # 如果不是数字字符串，则删除最后一行
    df = df.drop(df.index[-1])

# 定义一个函数来替换字符串中的空格，同时保持 None 或 NaN 不变
def replace_spaces(cell):
    if cell is None or pd.isna(cell):
        return cell
    else:
        return str(cell).replace(' ', '')

# 应用函数到DataFrame的每一列
df = df.applymap(replace_spaces)


df['值班领导'] = df['值班领导'].str.replace(' ', '', regex=True)
df['值班日期'] = df['值班日期'].str.replace('月', '', regex=True)

df = df.fillna(method='ffill')
# df['24小时值班电话'] = df['24小时值班电话'].fillna(method='ffill')
# df['值班日期'] = df['值班日期'].fillna(method='ffill')

df['手机'] = pd.to_numeric(df['手机'], errors='coerce')
df['24小时值班电话'] = pd.to_numeric(df['24小时值班电话'], errors='coerce')
df['值班日期'] = pd.to_numeric(df['值班日期'], errors='coerce')

# 这里假设名字和电话号码之间没有其他字符，只有空格
df['name'] = df['值班领导'].apply(lambda x: x[:-11].strip())
# 使用正则表达式来提取后11位数字作为一个字段
df['phone'] = df['值班领导'].apply(lambda x: re.search(r'\d{11}$', x).group())

df.drop(columns='值班领导', inplace=True)

df = df.rename(columns={'name': '值班领导', 'phone': '值班领导电话', '值班人员': '值班组员', '手机': '电话', 'Unnamed: 7': '日', '值班日期': '月'})


# 自定义函数来拆分日期范围并创建新行
def split_date_range(row):
    date_range = row['日']
    if '日' in date_range:  # 如果日期范围使用'日'分隔，则使用'日'分割
        dates = date_range.split('日')
    else:
        dates = [date_range]  # 如果只有一个日期，则保持原样

    # 创建新行并返回行列表
    new_rows = []
    for date in dates:
        if date:  # 确保日期非空
            new_row = row.copy()  # 复制原始行
            new_row['日'] = date  # 设置新的日期范围列
            new_rows.append(new_row)
    return new_rows


# 使用apply函数处理DataFrame，并收集所有新行
new_rows = []
for _, row in df.iterrows():
    new_rows.extend(split_date_range(row))

# 将新行转换为DataFrame
df = pd.DataFrame(new_rows)
df['值班组名'] = df['值班领导'] + '组'

df['日'] = df['日'].astype(str).str.replace('\n', '')
df['值班日期日期格式'] = pd.to_datetime('2024-' + df['月'].astype(str) + '-' + df['日'].astype(str))
df['值班日期'] = df['值班日期日期格式'].dt.strftime('%Y-%m-%d')
df = df.drop(columns=['序号', '职务', '月', '日'])
df = df.sort_values(by='值班日期')
# 使用 pd.to_numeric 将文本列转换为整数
df['值班领导电话'] = pd.to_numeric(df['值班领导电话'])
df2 = df
# df2.to_excel(('/home/sf107/桌面/test.xlsx'), index=False)
groupName = ''
for index, row in df2.iterrows():
    if row['值班组名'] != groupName:
        groupName = row['值班组名']
        # 组员行
        newRow2 = pd.Series(
            [
                row['值班领导'],
                row['值班领导电话'],
                28039061,
                row['值班领导'],
                row['值班领导电话'],
                row['值班组名'],
                row['值班日期日期格式'],
                row['值班日期']
            ],
            index=df2.columns
        )
        df2 = df2.append(newRow2, ignore_index=True)
df2 = df2.sort_values(by='值班日期')
df6 = df2
df6 = df6.rename(columns={'值班组员': '值班人员', '电话': '手机'})
# df6 = df6.drop_duplicates(subset='值班人员')
df6['职务'] = df6.apply(lambda row: '值班领导' if row['值班领导'] == row['值班人员'] else '值班组员', axis=1)
df6['手机文本格式'] = df6['手机'].map(int).astype(str)
df6['签到情况'] = '未签到'
df6['星期'] = ''
df8 = df6
df8['交班日期'] = df8['值班日期日期格式'] + pd.DateOffset(days=1)
# 原输出表最后一行
lastGroup = df8.at[df.__len__() - 1, '值班组名']  # 原表最后一个值班组
lastDate = df8.at[df.__len__() - 1, '值班日期日期格式']  # 原表最后一个值班日期日期格式
print(lastGroup)
print(lastDate)
# 确定组名第一个的排序
groupsSeries = df8['值班组名']
groupsSeries = groupsSeries.dropna()  # 去掉空值
groupsArr = groupsSeries.tolist()
print(groupsArr)
# 正确的使用方式是
unique_list = []
for x in groupsArr:
    if x not in unique_list:
        unique_list.append(x)
print(unique_list)  # 输出 [1, 2, 3, 4, 5]

# 处理list排序
list1 = unique_list[:unique_list.index(lastGroup) + 1]
list2 = unique_list[-(len(unique_list) - unique_list.index(lastGroup) - 1):]
unique_listNew = unique_list if unique_list.index(lastGroup) == 15 else list2 + list1
print(unique_listNew)


# df8.to_excel('/home/sf107/桌面/55555555555.xlsx', index=False)
df9 = df8.drop_duplicates(subset='值班人员')
# df9.to_excel('/home/sf107/桌面/777.xlsx', index=False)
len1 = len(df8) - len(df9)
print(len1)
# 使用 pd.concat 和列表推导式来重复 df 10 次
df_list = [df9 for _ in range(35)]
df10 = pd.concat(df_list, ignore_index=True)

# 现在 df2 包含 1000 行数据，是 df 的 10 倍
df10 = df10.iloc[len1:]
# lastDate
groupName2 = lastGroup
for index, row in df10.iterrows():
    if row['值班组名'] == groupName2:
        pass
    else:
        groupName2 = row['值班组名']
        lastDate = lastDate + timedelta(days=1)
    df10.at[index, '值班日期日期格式'] = lastDate
    df10.at[index, '交班日期'] = lastDate + timedelta(days=1)
    df10.at[index, '值班日期'] = lastDate.strftime('%Y-%m-%d')
# df10.to_excel('/home/sf107/桌面/666.xlsx', index=False)
df_combined = pd.concat([df8, df10], ignore_index=True)
df_combined = df_combined.drop(columns=['星期'])
merged = pd.merge(df_combined, df11[['日期', '星期']], left_on='值班日期日期格式', right_on='日期', how='left', suffixes=('', '_update'))
merged = merged.drop(columns=['日期'])
merged.to_excel('/home/sf107/桌面/值班安排表处理/值班明细表111.xlsx', index=False)
rn2('/home/sf107/桌面/值班安排表处理/值班明细表111.xlsx', '/home/sf107/桌面/值班安排表处理/值班明细表111.xlsx', ['A'])  # ！！！！！！！
# merged.to_excel('/home/sf107/桌面/888.xlsx', index=False)
df7 = pd.DataFrame(columns=df6.columns)
for index2, row2 in df5.iterrows():
    for index3, row3 in df6.iterrows():
        if row2['值班领导'] == row3['值班领导']:
            row3['星期'] = row2['星期']
            row3['值班日期日期格式'] = row2['日期']
            row3['值班日期'] = row2['日期'].strftime("%Y-%m-%d")
            # df7 = df7._append(row3, ignore_index=True)
            df7 = pd.concat([df7,pd.DataFrame([row3])], ignore_index=True, sort=False)
df7['交班日期'] = df7['值班日期日期格式'] + pd.DateOffset(days=1)
df7 = df7[['值班日期'] + [col for col in df7.columns if col != '值班日期']]

df7.to_excel('/home/sf107/桌面/值班安排表处理/节假日值班明细表.xlsx', index=False)
rn2('/home/sf107/桌面/值班安排表处理/节假日值班明细表.xlsx', '/home/sf107/桌面/值班安排表处理/节假日值班明细表.xlsx', ['B'])  # ！！！！！！！


def handel_excel(dataframe, o_name):
    df1 = dataframe.groupby(['值班日期日期格式'])['值班组员'].agg(','.join).reset_index()
    dataframe = pd.merge(dataframe, df1[['值班日期日期格式', '值班组员']], on='值班日期日期格式', how='left')

    dataframe = dataframe[['值班日期日期格式'] + [col for col in dataframe.columns if col != '值班日期日期格式']]

    dataframe.to_excel(o_name, index=False)
    # 打开现有的工作簿
    wb = openpyxl.load_workbook(o_name)
    ws = wb.worksheets[0]
    ws.insert_rows(2)

    # 合并单元格，小计
    current_num = 3
    count = 0
    for k in range(len(dataframe)):
        if k == 0:
            continue
        elif dataframe.iat[k, 5] != dataframe.iat[k - 1, 5]:
            ws.merge_cells(f'D{current_num}:D{current_num + count}')
            ws.merge_cells(f'E{current_num}:E{current_num + count}')
            ws.merge_cells(f'A{current_num}:A{current_num + count}')
            ws.merge_cells(f'F{current_num}:F{current_num + count}')
            ws.merge_cells(f'G{current_num}:G{current_num + count}')
            ws.merge_cells(f'H{current_num}:H{current_num + count}')
            ws.merge_cells(f'I{current_num}:I{current_num + count}')
            current_num = current_num + count + 1
            count = 0
        elif k == len(dataframe) - 1:
            ws.merge_cells(f'D{current_num}:D{current_num + count + 1}')
            ws.merge_cells(f'E{current_num}:E{current_num + count + 1}')
            ws.merge_cells(f'A{current_num}:A{current_num + count + 1}')
            ws.merge_cells(f'F{current_num}:F{current_num + count + 1}')
            ws.merge_cells(f'G{current_num}:G{current_num + count + 1}')
            ws.merge_cells(f'H{current_num}:H{current_num + count + 1}')
            ws.merge_cells(f'I{current_num}:I{current_num + count + 1}')
        else:
            count += 1

    # 处理表头
    ws.merge_cells('B1:C1')
    ws['B1'] = '值班人员子表单'
    ws['B2'] = '值班组员'
    ws['C2'] = '电话'
    ws['I1'] = '值班人员'
    ws.merge_cells('A1:A2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('E1:E2')
    ws.merge_cells('F1:F2')
    ws.merge_cells('G1:G2')
    ws.merge_cells('H1:H2')
    ws.merge_cells('I1:I2')

    wb.save(o_name)

    # 替换重名
    rn2(o_name, o_name, ['B', 'I'])  # ！！！！！！！


handel_excel(df, output_name)
handel_excel(df2, output_name2)
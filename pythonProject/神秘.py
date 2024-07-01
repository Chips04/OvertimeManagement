import openpyxl
import pandas as pd
import random
import calendar
import datetime

































today = datetime.date.today()

# today = datetime.date(2024, 11, 16)

for i1 in range(today.month):
    # 打开现有的工作簿
    wb_origin = openpyxl.load_workbook('/home/sf107/桌面/2生活垃圾产生量记录表/模板/生活垃圾产生量xx月份记录表.xlsx')
    ws_origin = wb_origin.active
    # 结果的表格
    columns = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']
    df_final = pd.DataFrame(columns=columns)
    if i1 == today.month - 1:
        days_in_month = today.day - 1
    else:
        days_in_month = calendar.monthrange(2024, i1 + 1)[1]
    for i in range(days_in_month):
        newRow1 = pd.Series(
            [
                int(0),
                int(0),
                int(0),
                int(0),
                round(random.uniform(1.3, 2.5), 1),
                round(random.uniform(1.0, 2.6), 1),
                int(0),
                int(0),
                int(random.randint(14, 26)),
                round(random.uniform(1.5, 2.4), 1)
            ],
            index=columns
        )
        df_final = df_final.append(newRow1, ignore_index=True)
    # 填充数据
    BBB = 7
    for i2 in range(len(df_final)):
        for j2 in range(len(df_final.columns)):
            cell_value = df_final.iat[i2, j2]
            ws_origin.cell(i2 + BBB, j2 + 2, value=cell_value)
    month_value = '0' + str(i1 + 1) if i1 +1 < 10 else str(i1 + 1)
    ws_origin.cell(1, 1, value=f"  2024年  {month_value}月  上芬社区工作站   生活垃圾产生量记录表")
    wb_origin.save(f"/home/sf107/桌面/2生活垃圾产生量记录表/2024/生活垃圾产生量{month_value}月份记录表.xlsx")